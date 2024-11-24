import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def processar_arquivo(caminho_arquivo):
    # Inicializa listas para armazenar as informações extraídas
    bairros = []
    causas = []
    enderecos = []
    inicios = []
    fins = []

    # Verifica se o arquivo existe
    if not os.path.exists(caminho_arquivo):
        print(f"Arquivo não encontrado: {caminho_arquivo}. Pulando...")
        return pd.DataFrame({
            'Bairro': bairros,
            'Causa': causas,
            'Endereço': enderecos,
            'Início': inicios,
            'Fim': fins
        })  # Retorna DataFrame vazio

    # Lê o arquivo TXT
    with open(caminho_arquivo, 'r', encoding='utf-8') as file:
        linhas = file.readlines()

    # Variáveis temporárias para armazenar os dados atuais
    bairro = None
    causa = None

    # Percorre as linhas
    for i in range(len(linhas)):
        linha = linhas[i].strip()

        # Atualiza o bairro quando encontramos uma linha que começa com "Bairro :"
        if linha.startswith("Bairro :"):
            bairro = linha.split(":")[1].strip()

        # Atualiza a causa quando encontramos uma linha que começa com "Causa :"
        elif linha.startswith("Causa :"):
            causa = linha.split(":")[1].strip()

        # Identifica o endereço e verifica as linhas seguintes para as datas
        elif linha.startswith("End.:"):
            endereco = linha.split(":")[1].strip()

            # Verifica as próximas linhas para capturar "Inicio" e "Final"
            inicio_linha = linhas[i + 1].strip() if i + 1 < len(linhas) else None
            final_linha = linhas[i + 2].strip() if i + 2 < len(linhas) else None

            inicio = re.search(r"Inicio: ([\d/ :]+)", inicio_linha).group(
                1).strip() if inicio_linha and "Inicio:" in inicio_linha else "Data de início não encontrada"
            fim = re.search(r"Final: ([\d/ :]+)", final_linha).group(
                1).strip() if final_linha and "Final:" in final_linha else "Data de fim não encontrada"

            # Adiciona as informações nas listas
            bairros.append(bairro)
            causas.append(causa)
            enderecos.append(endereco)
            inicios.append(inicio)
            fins.append(fim)

    # Cria o DataFrame
    df = pd.DataFrame({
        'Bairro': bairros,
        'Causa': causas,
        'Endereço': enderecos,
        'Início': inicios,
        'Fim': fins
    })

    return df


def formatar_tabelas_excel(arquivo_excel):
    # Abre o arquivo Excel usando openpyxl
    wb = load_workbook(arquivo_excel)
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Gera um nome de tabela válido (sem espaços)
        table_name = f"Table_{sheet.replace(' ', '_')}"

        # Define o intervalo da tabela (de A1 até a última linha e coluna)
        tabela = Table(displayName=table_name, ref=ws.dimensions)

        # Define o estilo da tabela
        estilo = TableStyleInfo(
            name="TableStyleMedium9",  # Escolha um estilo de tabela do Excel
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tabela.tableStyleInfo = estilo

        # Adiciona a tabela à aba
        ws.add_table(tabela)

    # Salva as alterações no arquivo Excel
    wb.save(arquivo_excel)


# Processa os arquivos e cria DataFrames
arquivos = {
    "Santo Amaro": r'/home/mais/celesc/arquivos_txt/santo_amaro_da_imperatriz.txt',
    "Palhoça": r'/home/mais/celesc/arquivos_txt/palhoca.txt',
    "Florianópolis": r'/home/mais/celesc/arquivos_txt/florianopolis.txt',
    "Garopaba": r'/home/mais/celesc/arquivos_txt/garopaba.txt',
    "São José": r'/home/mais/celesc/arquivos_txt/sao_jose.txt',
    "Águas Mornas": r'/home/mais/celesc/arquivos_txt/aguas_mornas.txt',
    "São Pedro de Alcântara": r'/home/mais/celesc/arquivos_txt/sao_pedro_de_alcantara.txt',
}

# Cria o arquivo Excel
arquivo_saida = '~/scripts/manutencoes_formatadas.xlsx'
with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
    for aba, caminho in arquivos.items():
        df = processar_arquivo(caminho)
        df.to_excel(writer, sheet_name=aba, index=False)

# Formata as abas do Excel como tabelas
formatar_tabelas_excel(arquivo_saida)

print("Arquivo Excel com múltiplas abas formatadas como tabelas gerado com sucesso!")
